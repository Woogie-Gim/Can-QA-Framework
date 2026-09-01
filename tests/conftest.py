import datetime
import os

import pytest
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# 리포트 출력 경로
REPORT_DIR = "reports"

# 테스트 함수명 대 검증 항목 설명
TEST_DESCRIPTIONS = {
    "test_message_received": "VehicleStatus 메시지 수신 여부",
    "test_cycle_time": "송신 주기 규격 준수 (100±20ms)",
    "test_signal_range": "신호 값의 DBC 정의 범위 준수",
    "test_alive_counter_increments": "롤링 카운터 연속성",
}

_results = []


def pytest_addoption(parser):
    parser.addoption(
        "--fault",
        action="store",
        default="none",
        help="리포트에 기록할 결함 주입 조건",
    )


@pytest.hookimpl(hookwrapper=True)
def pytest_runtest_makereport(item, call):
    outcome = yield
    report = outcome.get_result()
    # 실제 테스트 실행 단계만 수집
    if report.when != "call":
        return

    base_name = item.originalname or item.name
    _results.append(
        {
            "name": item.name,
            "description": TEST_DESCRIPTIONS.get(base_name, ""),
            "result": _resolve_result(report),
            "duration": report.duration,
            "message": _extract_reason(report),
        }
    )

def _resolve_result(report) -> str:
    # 검증 불가(SKIP)와 검증 실패(FAIL)를 구분
    if report.skipped:
        return "SKIP"
    if report.passed:
        return "PASS"
    return "FAIL"

def _extract_reason(report) -> str:
    # 실패 또는 건너뜀 사유 추출
    if report.passed:
        return ""
    if report.skipped:
        return "프레임 미수신으로 검증 불가"
    text = str(report.longrepr)
    for line in text.splitlines():
        if "AssertionError:" in line:
            return line.split("AssertionError:", 1)[1].strip()
    return "실패 사유 확인 필요"


def pytest_sessionfinish(session):
    if not _results:
        return

    fault = session.config.getoption("--fault")
    timestamp = datetime.datetime.now()

    wb = Workbook()
    ws = wb.active
    ws.title = "검증 결과"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4472C4")
    pass_fill = PatternFill("solid", fgColor="C6EFCE")
    fail_fill = PatternFill("solid", fgColor="FFC7CE")
    skip_fill = PatternFill("solid", fgColor="D9D9D9")

    # 요약 정보
    total = len(_results)
    failed = sum(1 for r in _results if r["result"] == "FAIL")
    skipped = sum(1 for r in _results if r["result"] == "SKIP")
    summary = [
        ("실행 일시", timestamp.strftime("%Y-%m-%d %H:%M:%S")),
        ("주입 결함", fault),
        ("전체 항목", total),
        ("성공", total - failed - skipped),
        ("실패", failed),
        ("검증 불가", skipped),
    ]
    
    for row, (key, value) in enumerate(summary, start=1):
        ws.cell(row=row, column=1, value=key).font = Font(bold=True)
        ws.cell(row=row, column=2, value=value)

    # 결과 표 헤더
    start_row = len(summary) + 2
    headers = ["No", "테스트 항목", "검증 내용", "결과", "소요(s)", "실패 사유"]
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=start_row, column=col, value=name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    # 결과 행
    for idx, item in enumerate(_results, start=1):
        row = start_row + idx
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=item["name"])
        ws.cell(row=row, column=3, value=item["description"])
        result_cell = ws.cell(row=row, column=4, value=item["result"])
        result_cell.alignment = Alignment(horizontal="center")
        fills = {"PASS": pass_fill, "FAIL": fail_fill, "SKIP": skip_fill}
        result_cell.fill = fills[item["result"]]
        ws.cell(row=row, column=5, value=round(item["duration"], 4))
        ws.cell(row=row, column=6, value=item["message"])

    widths = [18, 40, 32, 10, 10, 50]
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    os.makedirs(REPORT_DIR, exist_ok=True)
    path = os.path.join(
        REPORT_DIR,
        f"report_{fault}_{timestamp.strftime('%Y%m%d_%H%M%S')}.xlsx",
    )
    wb.save(path)
    print(f"\n리포트 생성: {path}")
